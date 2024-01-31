import tempfile
import logging
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, Border, PatternFill, Alignment, Side

from spaceone.core.manager import BaseManager
from spaceone.inventory.manager.file_manager import FileManager
from spaceone.inventory.error.export import *

_LOGGER = logging.getLogger(__name__)


class ExportManager(BaseManager):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self._file_format = kwargs.get("file_format", "EXCEL")

        file_name = kwargs.get("file_name", "export")
        now = datetime.utcnow()

        if self._file_format == "EXCEL":
            self._file_name = f'{file_name}_{now.strftime("%Y%m%d")}.xlsx'
        else:
            self._file_name = f'{file_name}_{now.strftime("%Y%m%d")}.zip'

        self._file_dir = None
        self._file_path = None
        self._sheet_name_count = {}

    def export(
        self, export_options: dict, domain_id: str, workspace_id: str = None, **kwargs
    ) -> dict:
        with tempfile.TemporaryDirectory() as temp_dir:
            self._file_dir = temp_dir
            self._file_path = f"{self._file_dir}/{self._file_name}"
            self.make_file(export_options)
            return self.upload_file(domain_id, workspace_id)

    def make_file(self, export_options: dict) -> None:
        self._check_results(export_options)

        if self._file_format == "EXCEL":
            with pd.ExcelWriter(self._file_path, mode="w", engine="openpyxl") as writer:
                idx = 0

                for export_option in export_options:
                    name = export_option["name"]
                    title = export_option.get("title")
                    results = export_option["results"]

                    if len(results) > 0:
                        self._make_excel_file(writer, idx, name, results, title)

                        idx += 1

        else:
            raise ERROR_NOT_SUPPORT_FILE_FORMAT(file_format=self._file_format)

    @staticmethod
    def _change_sheet_name(name: str) -> str:
        return (
            name.replace(" ", "")
            .replace("/", "")
            .replace("\\", "")
            .replace("?", "")
            .replace("*", "")
            .replace("[", "")
            .replace("]", "")
            .replace(":", "")
        )[:30]

    @staticmethod
    def _get_default_font(is_header: bool = False) -> Font:
        return Font(size=12, bold=is_header, color="FFFFFF" if is_header else "000000")

    @staticmethod
    def _write_excel_file(
        writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, title: str = None
    ) -> None:
        start_row = 1 if title else 0

        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
        ws = writer.sheets[sheet_name]

        # Set Title
        if title:
            ws = writer.sheets[sheet_name]
            ws["A1"] = title
            ws["A1"].font = Font(size=24, bold=True, color="0A3763")

        # Set Excel Style
        # Header Style
        align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_border = Border(
            right=Side(style="thin", color="FFFFFF"), left=None, top=None, bottom=None
        )
        header_fill = PatternFill(patternType="solid", fgColor="0A3764")

        # Data Style
        data_font = Font(size=12, bold=False, color="000000")
        data_border = Border(
            left=Side(style="thin", color="E9E9EC"),
            top=Side(style="thin", color="E9E9EC"),
            right=Side(style="thin", color="E9E9EC"),
            bottom=Side(style="thin", color="E9E9EC"),
        )
        data_fill = PatternFill(patternType="solid", fgColor="F7f7f7")

        for col in ws.columns:
            max_width = 0
            for i, cell in enumerate(col):
                if i >= start_row:
                    if i == start_row:
                        cell.alignment = align
                        cell.font = header_font
                        cell.border = header_border
                        cell.fill = header_fill
                    else:
                        cell.alignment = align
                        cell.font = data_font
                        cell.border = data_border

                        if i % 2 == 0:
                            cell.fill = data_fill

                    if len(str(cell.value)) > max_width:
                        for x in str(cell.value).split("\n"):
                            if len(x) > max_width:
                                max_width = len(x)

                ws.column_dimensions[col[0].column_letter].width = (max_width + 2) * 1.1

    def _make_excel_file(
        self,
        writer: pd.ExcelWriter,
        idx: int,
        name: str,
        results: list,
        title: str = None,
    ) -> None:
        df = pd.DataFrame(results)

        sheet_name = self._change_sheet_name(name)

        if sheet_name in self._sheet_name_count:
            self._sheet_name_count[sheet_name] += 1
            count = self._sheet_name_count[sheet_name]

            sheet_name = f"{sheet_name[:29]}{count}"

        else:
            self._sheet_name_count[sheet_name] = 1

        if idx == 0:
            self._write_excel_file(writer, df, sheet_name, title)
        else:
            self._write_excel_file(writer, df, sheet_name, title)

    def upload_file(self, domain_id: str, workspace_id: str = None) -> dict:
        file_mgr: FileManager = self.locator.get_manager(FileManager)

        params = {
            "name": self._file_name,
            "domain_id": domain_id,
            "resource_group": "DOMAIN",
        }

        if workspace_id:
            role_type = self.transaction.get_meta("role_type")
            params["workspace_id"] = workspace_id

            if role_type == "WORKSPACE_OWNER":
                params["resource_group"] = "WORKSPACE"
            else:
                params["resource_group"] = "PROJECT"

        file_info = file_mgr.add_file(params, domain_id)

        file_mgr.upload_file(
            self._file_path, file_info["upload_url"], file_info["upload_options"]
        )
        download_file_info = file_mgr.get_download_url(file_info["file_id"], domain_id)

        return {"download_url": download_file_info["download_url"]}

    @staticmethod
    def _check_results(export_options: dict) -> None:
        has_results = False

        for export_option in export_options:
            results = export_option["results"]
            if len(results) > 0:
                has_results = True
                break

        if not has_results:
            raise ERROR_NO_DATA_TO_EXPORT()
